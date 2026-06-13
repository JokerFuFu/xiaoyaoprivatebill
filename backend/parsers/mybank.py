"""
浙江网商银行解析器 —— 网商账户收支明细 + 余利宝资金明细(两种 xlsx 同一家)。

为什么单独做:这俩是支付宝生态的「钱包/货基」账户,导出 xlsx,既不是微信也不是标准银行 PDF。
价值:① 余利宝「正常收益」是真实投资收益(此前账上几乎看不到);② 两个账户的余额列可看资产。

口径(关键,防与支付宝/微信流水双算):
- 凡是「快捷支付-支付宝/微信」「对方机构=支付宝/财付通」的流出 → 平台代扣(不计收支),
  真正的消费明细在支付宝/微信账单里,这里只是出资腿,绝不重复计消费。
- 余利宝 正常收益/利息 → 收入·投资收益(全新,不在别处)。
- 余利宝 申购/赎回、网商↔余利宝/自转 → 转入/转出·理财投资(内部搬运)。
- 其余按金额正负兜底为 转入/转出(对外往来),不轻易记消费(网商卡极少直接刷卡)。
"""
import logging
import re

import pandas as pd
import openpyxl

logger = logging.getLogger(__name__)

_PLATFORM_HINT = ('支付宝', '财付通', '快捷支付-支付宝', '快捷支付-微信', '微信支付')
# 真实收入的精确短语 —— 不能用裸"收益"(会误中"固定收益类理财"这种产品名)
_INCOME_RE = re.compile(r'正常收益|收益发放|结息|应付利息|派息|分红到账')
# 理财/货基 申赎、账户互转(内部搬运,非收支)
_FUND_RE = re.compile(r'申购|赎回|转入|转出|周利宝|余利宝|理财|货币基金|基金|蚂蚁财富|撤销|定期|存单')


def _open_rows(filepath):
    """读全部行。网商导出的 xlsx <dimension> 标签谎报行数(只声明到 11 行),
    read_only 模式会信它而漏读 → 必须 reset_dimensions() 强制全表扫描。"""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.worksheets[0]
    ws.reset_dimensions()
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


def _peek_title(filepath):
    """读首格判断是不是网商系 xlsx,以及是「收支」还是「余利宝」。"""
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb.worksheets[0]
        ws.reset_dimensions()
        head = ''
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            head += ' '.join(str(c) for c in row if c is not None)
            if i >= 8:
                break
        wb.close()
        return head
    except Exception:
        return ''


def is_mybank_xlsx(filepath):
    head = _peek_title(filepath)
    return ('网商银行' in head) or ('余利宝' in head) or ('浙江网商' in head)


def _find_header(ws_rows):
    for i, row in enumerate(ws_rows):
        cells = [str(c).strip() if c is not None else '' for c in row]
        if any(c.startswith('交易时间') for c in cells):
            return i, cells
    return None, None


def _col(header, *names):
    """在表头里找列索引(表头可能带换行/空白)。"""
    norm = [h.replace('\n', '').strip() for h in header]
    for nm in names:
        for j, h in enumerate(norm):
            if h.startswith(nm):
                return j
    return None


def _num(v):
    try:
        return float(str(v).replace(',', '').replace('+', '').replace('¥', '').strip() or 0)
    except (TypeError, ValueError):
        return None


def parse_mybank_xlsx(filepath):
    """解析网商/余利宝 xlsx → 标准 schema DataFrame,来源='银行'。"""
    head = _peek_title(filepath)
    is_yulibao = '余利宝' in head and '账户收支' not in head
    rows = _open_rows(filepath)

    hi, header = _find_header(rows)
    if hi is None:
        raise ValueError('网商 xlsx 未找到表头行')

    c_time = _col(header, '交易时间')
    c_amt = _col(header, '交易金额')
    c_party = _col(header, '对方户名')
    c_org = _col(header, '对方机构名称', '对方机构')
    c_type = _col(header, '交易类型')          # 仅余利宝有
    c_name = _col(header, '交易名称')          # 仅收支有
    c_memo = _col(header, '备注')              # 余利宝有

    pay = '余利宝' if is_yulibao else '网商银行(3087)'
    recs = []
    for row in rows[hi + 1:]:
        if row is None or c_time is None or c_time >= len(row):
            continue
        t = row[c_time]
        if t is None or not str(t).strip():
            continue
        amt = _num(row[c_amt]) if c_amt is not None and c_amt < len(row) else None
        if amt is None:
            continue
        party = str(row[c_party]).strip() if c_party is not None and c_party < len(row) and row[c_party] else ''
        org = str(row[c_org]).strip() if c_org is not None and c_org < len(row) and row[c_org] else ''
        ttype = str(row[c_type]).strip() if c_type is not None and c_type < len(row) and row[c_type] else ''
        tname = str(row[c_name]).strip() if c_name is not None and c_name < len(row) and row[c_name] else ''
        memo = str(row[c_memo]).strip() if c_memo is not None and c_memo < len(row) and row[c_memo] else ''
        text = ' '.join([ttype, tname, memo, party, org])

        is_platform = any(k in text for k in _PLATFORM_HINT)
        is_income = bool(_INCOME_RE.search(text)) and amt > 0
        is_fund = bool(_FUND_RE.search(text))

        # 优先级:真实收益(精确短语) > 平台代扣(防双算) > 理财申赎/搬运 > 按金额兜底
        if is_income:
            zhi, cat = '收入', '投资收益'                       # 利息/收益发放(真实投资收益,不在别处)
        elif is_platform:
            zhi, cat = '不计收支', '平台代扣'                   # 出资腿,真正消费在支付宝/微信账单
        elif is_fund:
            zhi, cat = ('转入' if amt > 0 else '转出'), '理财投资'   # 申赎/搬运(内部)
        elif amt > 0:
            zhi, cat = '转入', '资金往来'
        else:
            zhi, cat = '转出', '对外转账'

        recs.append({
            '交易时间': str(t).strip(),
            '交易分类': cat,
            '交易对方': party or org,
            '商品说明': (tname or memo or ttype)[:60],
            '收/支': zhi,
            '金额': abs(amt),
            '收/付款方式': pay,
            '交易状态': '交易成功',
            '来源': '银行',
            '是否退款': False,
        })

    df = pd.DataFrame(recs)
    if df.empty:
        return df
    df['交易时间'] = pd.to_datetime(df['交易时间'], errors='coerce')
    df = df[df['交易时间'].notna()].copy()
    df['月份'] = df['交易时间'].dt.strftime('%Y-%m')
    df['日期'] = df['交易时间'].dt.strftime('%Y-%m-%d')
    logger.info(f"解析网商{'(余利宝)' if is_yulibao else '(收支)'} {len(df)} 条")
    return df


def latest_balance(filepath):
    """读最新一条的余额(供资产快照参考),返回 (账户名, 余额) 或 None。"""
    try:
        head = _peek_title(filepath)
        is_yulibao = '余利宝' in head and '账户收支' not in head
        rows = _open_rows(filepath)
        hi, header = _find_header(rows)
        if hi is None:
            return None
        c_time = _col(header, '交易时间')
        c_bal = _col(header, '余额')
        if c_bal is None:
            return None
        best_t, best_v = None, None
        for row in rows[hi + 1:]:
            if not row or c_time >= len(row) or row[c_time] is None:
                continue
            t = pd.to_datetime(str(row[c_time]).strip(), errors='coerce')
            v = _num(row[c_bal]) if c_bal < len(row) else None
            if t is not None and not pd.isna(t) and v is not None:
                if best_t is None or t > best_t:
                    best_t, best_v = t, v
        if best_v is None:
            return None
        return ('余利宝' if is_yulibao else '网商银行(3087)', round(best_v, 2), best_t.strftime('%Y-%m-%d'))
    except Exception:
        logger.exception("读取网商余额失败")
        return None
